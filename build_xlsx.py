#!/usr/bin/env python3
"""
Step 3 (optional): Build an .xlsx with each card's screenshot embedded in-row.

Reads output/gift_cards.csv + the screenshots and writes output/gift_cards.xlsx.
Open it in Excel/Numbers (images show inline) or upload to Google Drive.

    python build_xlsx.py
"""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

BASE = Path(__file__).parent
CARDS_CSV = BASE / "output" / "gift_cards.csv"
XLSX = BASE / "output" / "gift_cards.xlsx"
THUMBS = BASE / "screenshots" / "thumbs"

# Crop the screenshot to the card area (full page is mostly whitespace), then
# scale to this width. Tune CROP if the retailer's layout differs.
CROP = (0, 0, 470, 860)   # left, top, right, bottom in screenshot pixels
THUMB_W = 200             # thumbnail width in pixels

COLUMNS = ["index", "code", "pin", "amount", "from", "card", "date", "subject"]


def make_thumb(src: Path, dst: Path) -> tuple[int, int]:
    img = PILImage.open(src)
    box = (CROP[0], CROP[1], min(CROP[2], img.width), min(CROP[3], img.height))
    img = img.crop(box)
    h = round(img.height * (THUMB_W / img.width))
    img = img.resize((THUMB_W, h))
    dst.parent.mkdir(parents=True, exist_ok=True)
    img.save(dst)
    return THUMB_W, h


def main():
    rows = list(csv.DictReader(open(CARDS_CSV)))
    wb = Workbook()
    ws = wb.active
    ws.title = "Gift Cards"

    # Header
    for c, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    img_col = COLUMNS.index("card") + 1
    img_col_letter = get_column_letter(img_col)

    for i, r in enumerate(rows, start=2):
        # text columns (strip the spreadsheet-safety apostrophe; xlsx text cells
        # preserve leading zeros natively)
        ws.cell(row=i, column=1, value=int(r["index"]))
        for name in ["code", "pin", "amount", "from", "date", "subject"]:
            col = COLUMNS.index(name) + 1
            cell = ws.cell(row=i, column=col, value=r[name].lstrip("'"))
            if name in ("code", "pin"):
                cell.number_format = "@"  # force text

        # embedded screenshot
        shot = BASE / r["screenshot"]
        if shot.exists():
            thumb = THUMBS / f"{int(r['index']):03}.png"
            w, h = make_thumb(shot, thumb)
            xi = XLImage(str(thumb))
            ws.add_image(xi, f"{img_col_letter}{i}")
            ws.row_dimensions[i].height = h * 0.75  # px -> points

    # column widths
    widths = {"index": 7, "code": 22, "pin": 12, "amount": 9,
              "from": 18, "card": THUMB_W / 7, "date": 22, "subject": 28}
    for name, w in widths.items():
        ws.column_dimensions[get_column_letter(COLUMNS.index(name) + 1)].width = w
    for col in ["A", "B", "C", "D", "E"]:
        for cell in ws[col]:
            cell.alignment = Alignment(vertical="top")

    wb.save(XLSX)
    print(f"Wrote {XLSX} with {len(rows)} cards (images embedded).")


if __name__ == "__main__":
    main()
